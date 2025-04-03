import openpyxl
from openpyxl.styles.fills import PatternFill

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum,columnno).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum,columnno).value=data
    workbook.save(file)

def fillGreenColor(file,sheetName,rownum,columnno):
    workboook = openpyxl.load_workbook(file)
    sheet = workboook[sheetName]
    greenFill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columnno).fill = greenFill
    workboook.save(file)

def fillRedcolor(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,columnno).fill=redFill
    workbook.save(file)