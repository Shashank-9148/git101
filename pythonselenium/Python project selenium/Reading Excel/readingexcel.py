import openpyxl

#file--> workbook --> sheets -->Rows -->columns
file = "C:\\Users\\Shashank L\\Downloads\\Book1.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row #count number of rows in a excel sheet 4
cols = sheet.max_column #count number of columns in a excel sheet 4

#reading all the rows and columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="      ")
    print()