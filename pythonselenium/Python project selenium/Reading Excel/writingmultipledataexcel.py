import openpyxl
from openpyxl.cell import cell

file = r"C:\Users\Shashank L\Downloads\write.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active #or sheet = workbook["Data"] -->get active sheet from excel

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = "Smith"
sheet.cell(1,3).value = "Engineer"

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = "John"
sheet.cell(2,3).value = "Manager"

sheet.cell(3,1).value = 890
sheet.cell(3,2).value = "David"
sheet.cell(3,3).value = "Developer"

workbook.save(file)